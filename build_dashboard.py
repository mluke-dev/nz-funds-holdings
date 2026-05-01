#!/usr/bin/env python3
"""
Generate a static HTML dashboard from the cleaned holdings data.

v2 improvements:
  - Heat-map coloring (darker = bigger position)
  - Toggle button: weights ↔ active vs NZX50 benchmark
  - "Held by" concentration column
  - Tooltips on long fund names
  - Sticky first column

Reads:  data/holdings_clean.csv, data/holdings_matrix.xlsx, selected_funds.csv
Writes: data/dashboard.html
"""

from __future__ import annotations
import csv
import os
from datetime import datetime
from openpyxl import load_workbook

SCRIPT_DIR    = os.path.dirname(os.path.abspath(__file__))
DATA_DIR      = os.path.join(SCRIPT_DIR, "data")
CLEAN_CSV     = os.path.join(DATA_DIR, "holdings_clean.csv")
MATRIX_XLSX   = os.path.join(DATA_DIR, "holdings_matrix.xlsx")
SELECTED_CSV  = os.path.join(SCRIPT_DIR, "selected_funds.csv")
DASHBOARD_HTML= os.path.join(DATA_DIR, "dashboard.html")


def load_clean_rows():
    rows = []
    with open(CLEAN_CSV, encoding="utf-8") as f:
        for r in csv.DictReader(f):
            try:
                r["weight_pct"] = float(r["weight_pct"])
            except (KeyError, ValueError):
                r["weight_pct"] = 0.0
            rows.append(r)
    return rows


def load_fund_meta():
    out = {}
    with open(SELECTED_CSV, encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            out[r["fund_id"]] = {
                "fund_name":       r.get("fund_name", ""),
                "provider":        r.get("provider", ""),
                "fees_pct":        r.get("fees_pct", ""),
                "total_value_nzd": r.get("total_value_nzd", ""),
            }
    return out


def load_benchmark_from_matrix() -> dict:
    if not os.path.exists(MATRIX_XLSX):
        return {}
    wb = load_workbook(MATRIX_XLSX, data_only=True)
    if "Matrix" not in wb.sheetnames:
        return {}
    ws = wb["Matrix"]
    bench_col = None
    for c in range(1, ws.max_column + 1):
        if ws.cell(1, c).value == "BENCHMARK":
            bench_col = c
            break
    if not bench_col:
        return {}
    out = {}
    for r in range(2, ws.max_row + 1):
        t = ws.cell(r, 1).value
        w = ws.cell(r, bench_col).value
        if t and isinstance(w, (int, float)):
            out[t] = w
    return out


def build_pivot(rows):
    pivot, meta, fund_ids = {}, {}, set()
    for r in rows:
        t = r["ticker"]; f = r["fund_id"]
        fund_ids.add(f)
        pivot.setdefault(t, {})[f] = pivot.get(t, {}).get(f, 0.0) + r["weight_pct"]
        meta[t] = {
            "name":    r["canonical_name"],
            "country": r["country"],
            "sector":  r["sector"],
        }
    return pivot, meta, fund_ids


HTML_TEMPLATE = r"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<title>NZ Australasian Funds — Holdings Dashboard</title>
<style>
  :root {
    --bg: #f5f7fa; --panel: #fff; --ink: #1a2332; --muted: #6b7585;
    --border: #e1e5ec; --accent: #2c3e50;
    --pos: #1e7c4a; --neg: #c0392b;
    --bench: #16a085; --bench-bg: #e8f5f1;
    --row-alt: #fafbfc;
    --w1: #ecf3fb; --w2: #d4e6f8; --w3: #a8caee; --w4: #5e9ad8; --w5: #2c3e50;
  }
  * { box-sizing: border-box; }
  body {
    margin: 0; background: var(--bg); color: var(--ink);
    font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
    font-size: 13px; line-height: 1.4;
  }
  header {
    background: var(--accent); color: #fff; padding: 16px 24px;
    display: flex; justify-content: space-between; align-items: center;
    flex-wrap: wrap; gap: 12px;
  }
  header h1 { font-size: 16px; font-weight: 600; margin: 0; }
  header .meta { font-size: 11px; opacity: 0.75; margin-top: 2px; }

  .toolbar {
    display: flex; gap: 12px; padding: 12px 24px; background: #fff;
    border-bottom: 1px solid var(--border); flex-wrap: wrap; align-items: center;
  }
  .toolbar label {
    font-size: 12px; color: var(--muted);
    display: flex; align-items: center; gap: 4px;
  }
  .toolbar input[type="text"], .toolbar select {
    border: 1px solid var(--border); border-radius: 4px;
    padding: 5px 8px; font-size: 13px; background: #fff;
  }
  .toolbar input[type="text"] { width: 200px; }
  .spacer { flex: 1; }

  .toggle-group {
    display: inline-flex; border: 1px solid var(--border); border-radius: 4px;
    overflow: hidden;
  }
  .toggle-group button {
    border: none; background: #fff; padding: 6px 12px; font-size: 12px;
    cursor: pointer; color: var(--muted); font-weight: 500;
  }
  .toggle-group button:hover { background: #f0f3f7; }
  .toggle-group button.active { background: var(--accent); color: #fff; }

  .table-wrap {
    margin: 0 24px 24px;
    background: var(--panel); border: 1px solid var(--border);
    border-radius: 6px; overflow: auto;
    max-height: calc(100vh - 130px);
  }
  table { border-collapse: collapse; width: 100%; font-size: 12px; }

  thead th {
    position: sticky; top: 0; z-index: 3;
    background: var(--accent); color: #fff;
    text-align: left; padding: 8px 8px;
    font-weight: 600; font-size: 10.5px;
    text-transform: uppercase; letter-spacing: 0.4px;
    cursor: pointer; user-select: none;
    white-space: nowrap;
    border-right: 1px solid rgba(255,255,255,0.08);
  }
  thead th:hover { background: #34495e; }
  thead th.bench { background: var(--bench); }
  thead th.bench:hover { background: #138a72; }
  thead th .arrow { opacity: 0.35; margin-left: 4px; font-size: 9px; }
  thead th.sorted .arrow { opacity: 1; }
  thead th .fid {
    display: block; font-size: 9.5px; opacity: 0.6; margin-top: 2px;
    text-transform: none; letter-spacing: normal; font-weight: 400;
  }

  tbody td {
    padding: 6px 8px; border-bottom: 1px solid var(--border);
    text-align: right; white-space: nowrap;
  }
  tbody tr:nth-child(even) { background: var(--row-alt); }
  tbody tr:hover { background: #eef3f8 !important; }
  tbody td.text { text-align: left; }
  tbody td.ticker {
    font-weight: 600; font-family: ui-monospace, monospace; font-size: 11.5px;
    position: sticky; left: 0; background: inherit; z-index: 1;
  }
  tbody tr:nth-child(even) td.ticker { background: var(--row-alt); }
  tbody tr:hover td.ticker { background: #eef3f8 !important; }
  tbody td.country { font-size: 10.5px; color: var(--muted); text-align: center; width: 32px; }
  tbody td.sector  { font-size: 11.5px; color: var(--muted); }
  tbody td.held-by {
    font-size: 10.5px; text-align: center; color: var(--muted);
    background: #f7f9fb;
  }
  tbody td.held-by strong { color: var(--ink); }

  tbody td.w {
    font-feature-settings: "tnum"; font-variant-numeric: tabular-nums;
  }
  tbody td.w.zero { color: #d4dae3; }

  /* Heat tones — weight mode */
  tbody td.w.h1 { background: var(--w1); }
  tbody td.w.h2 { background: var(--w2); }
  tbody td.w.h3 { background: var(--w3); }
  tbody td.w.h4 { background: var(--w4); color: #fff; }
  tbody td.w.h5 { background: var(--w5); color: #fff; font-weight: 600; }

  /* Active mode tones */
  body.active-mode tbody td.w { background: transparent !important; color: var(--ink); }
  body.active-mode tbody td.w.zero { color: #d4dae3; }
  body.active-mode tbody td.w.act-pos1 { background: rgba(30,124,74,0.10); color: var(--pos); }
  body.active-mode tbody td.w.act-pos2 { background: rgba(30,124,74,0.22); color: var(--pos); font-weight: 600; }
  body.active-mode tbody td.w.act-pos3 { background: rgba(30,124,74,0.38); color: #145a36; font-weight: 600; }
  body.active-mode tbody td.w.act-neg1 { background: rgba(192,57,43,0.10); color: var(--neg); }
  body.active-mode tbody td.w.act-neg2 { background: rgba(192,57,43,0.22); color: var(--neg); font-weight: 600; }
  body.active-mode tbody td.w.act-neg3 { background: rgba(192,57,43,0.38); color: #8b1e15; font-weight: 600; }

  thead th.bench, tbody td.bench { border-left: 2px solid var(--bench); }
  tbody td.bench { background: var(--bench-bg); font-weight: 500; }

  footer { text-align: center; color: var(--muted); padding: 12px; font-size: 11px; }

  .legend {
    display: inline-flex; gap: 4px; align-items: center;
    font-size: 11px; color: var(--muted);
  }
  .legend .swatch {
    display: inline-block; width: 14px; height: 14px; border-radius: 2px;
    border: 1px solid rgba(0,0,0,0.05);
  }
  body.active-mode .legend.weight { display: none; }
  body:not(.active-mode) .legend.active { display: none; }
</style>
</head>
<body>

<header>
  <div>
    <h1>NZ Australasian Funds — Holdings Dashboard</h1>
    <div class="meta">__GENERATED__ · __FUND_COUNT__ funds · __TICKER_COUNT__ securities · __HOLDINGS_COUNT__ holdings</div>
  </div>
  <div class="toggle-group">
    <button data-mode="weight" class="active" id="modeWeight">Weights</button>
    <button data-mode="active" id="modeActive">Active vs NZX50</button>
  </div>
</header>

<section class="toolbar">
  <label>Search: <input type="text" id="filter" placeholder="Ticker, name or sector"></label>
  <label>Sector: <select id="sectorFilter"><option value="">All</option>__SECTORS__</select></label>
  <label>Country: <select id="countryFilter">
    <option value="">All</option><option value="NZ">NZ only</option><option value="AU">AU only</option>
  </select></label>
  <label>Held by: <select id="minFundsFilter">
    <option value="0">Any</option>
    <option value="2">≥ 2 funds</option>
    <option value="5">≥ 5 funds</option>
    <option value="10">≥ 10 funds</option>
  </select></label>
  <span class="spacer"></span>
  <span class="legend weight">
    Heat:
    <span class="swatch" style="background: var(--w1)"></span>
    <span class="swatch" style="background: var(--w2)"></span>
    <span class="swatch" style="background: var(--w3)"></span>
    <span class="swatch" style="background: var(--w4)"></span>
    <span class="swatch" style="background: var(--w5)"></span>
    larger position →
  </span>
  <span class="legend active">
    <span class="swatch" style="background: rgba(192,57,43,0.38)"></span> underweight
    <span class="swatch" style="background: rgba(30,124,74,0.38); margin-left:6px"></span> overweight
  </span>
</section>

<div class="table-wrap">
  <table id="matrix">
    <thead>__THEAD__</thead>
    <tbody>__TBODY__</tbody>
  </table>
</div>

<footer>
  Source: NZ Disclose Register · Active = fund weight − NZX50 benchmark · Click headers to sort
</footer>

<script>
(function () {
  const tbl = document.getElementById('matrix');
  const headers = tbl.querySelectorAll('thead th');
  let sortCol = -1, sortAsc = false;

  headers.forEach((th, i) => {
    th.addEventListener('click', () => {
      const sortType = th.dataset.sort || 'string';
      if (sortCol === i) sortAsc = !sortAsc; else { sortCol = i; sortAsc = false; }
      headers.forEach(h => {
        h.classList.remove('sorted');
        const a = h.querySelector('.arrow');
        if (a) a.textContent = '⇅';
      });
      th.classList.add('sorted');
      const arrow = th.querySelector('.arrow');
      if (arrow) arrow.textContent = sortAsc ? '▲' : '▼';

      const tbody = tbl.tBodies[0];
      const rows = Array.from(tbody.rows);
      const isActive = document.body.classList.contains('active-mode');
      rows.sort((a, b) => {
        const cellA = a.cells[i], cellB = b.cells[i];
        if (sortType === 'number') {
          let av, bv;
          if (isActive && cellA.dataset.active !== undefined && cellB.dataset.active !== undefined) {
            av = parseFloat(cellA.dataset.active) || 0;
            bv = parseFloat(cellB.dataset.active) || 0;
          } else {
            av = parseFloat(cellA.dataset.v) || 0;
            bv = parseFloat(cellB.dataset.v) || 0;
          }
          return sortAsc ? av - bv : bv - av;
        }
        const av = cellA.dataset.v || cellA.innerText.trim();
        const bv = cellB.dataset.v || cellB.innerText.trim();
        return sortAsc ? av.localeCompare(bv) : bv.localeCompare(av);
      });
      rows.forEach(r => tbody.appendChild(r));
    });
  });

  // Filtering
  const filterInput = document.getElementById('filter');
  const sectorSel   = document.getElementById('sectorFilter');
  const countrySel  = document.getElementById('countryFilter');
  const minFundsSel = document.getElementById('minFundsFilter');

  function applyFilters() {
    const text = (filterInput.value || '').toLowerCase();
    const sector = sectorSel.value;
    const country = countrySel.value;
    const minFunds = parseInt(minFundsSel.value, 10) || 0;
    Array.from(tbl.tBodies[0].rows).forEach(row => {
      const haystack = (row.dataset.searchable || '').toLowerCase();
      const rs = row.dataset.sector || '';
      const rc = row.dataset.country || '';
      const nf = parseInt(row.dataset.numFundsHolding, 10) || 0;
      let show = true;
      if (text && !haystack.includes(text)) show = false;
      if (sector && rs !== sector) show = false;
      if (country && rc !== country) show = false;
      if (nf < minFunds) show = false;
      row.style.display = show ? '' : 'none';
    });
  }
  [filterInput, sectorSel, countrySel, minFundsSel].forEach(el =>
    el.addEventListener('input', applyFilters)
  );

  // Mode toggle
  function setMode(mode) {
    document.querySelectorAll('.toggle-group button').forEach(b =>
      b.classList.toggle('active', b.dataset.mode === mode)
    );
    document.body.classList.toggle('active-mode', mode === 'active');
    const isActive = mode === 'active';
    document.querySelectorAll('tbody td.w').forEach(td => {
      const w = td.dataset.v;
      const a = td.dataset.active;
      // Don't transform the benchmark column itself
      if (td.classList.contains('bench')) {
        td.innerHTML = formatWeight(parseFloat(w));
        return;
      }
      if (isActive && a !== undefined) {
        td.innerHTML = formatActive(parseFloat(a));
      } else {
        td.innerHTML = formatWeight(parseFloat(w));
      }
    });
  }
  function formatWeight(w) {
    if (!w || Math.abs(w) < 0.005) return '<span style="color:#d4dae3">—</span>';
    return w.toFixed(2) + '%';
  }
  function formatActive(a) {
    if (a === undefined || a === null || isNaN(a) || Math.abs(a) < 0.005) {
      return '<span style="color:#d4dae3">—</span>';
    }
    return (a > 0 ? '+' : '') + a.toFixed(2) + '%';
  }
  document.querySelectorAll('.toggle-group button').forEach(b =>
    b.addEventListener('click', () => setMode(b.dataset.mode))
  );

  setMode('weight');
})();
</script>
</body>
</html>
"""


def render_dashboard():
    rows = load_clean_rows()
    fund_meta = load_fund_meta()
    bench = load_benchmark_from_matrix()
    pivot, meta, fund_ids_in_data = build_pivot(rows)

    fund_ids = [fid for fid in fund_meta if fid in fund_ids_in_data]
    for fid in sorted(fund_ids_in_data):
        if fid not in fund_ids:
            fund_ids.append(fid)

    sorted_tickers = sorted(pivot.keys(), key=lambda t: -sum(pivot[t].values()))
    total_funds = len(fund_ids)
    total_tickers = len(sorted_tickers)
    total_holdings = len(rows)
    sectors = sorted({meta[t]["sector"] for t in sorted_tickers if meta[t]["sector"]})

    # ── thead ──────────────────────────────────────────────────────────────
    thead_html = ['<tr>']
    thead_html.append('<th data-sort="string">Ticker<span class="arrow">⇅</span></th>')
    thead_html.append('<th data-sort="string">Name<span class="arrow">⇅</span></th>')
    thead_html.append('<th data-sort="string" style="text-align:center">Cty<span class="arrow">⇅</span></th>')
    thead_html.append('<th data-sort="string">Sector<span class="arrow">⇅</span></th>')
    thead_html.append('<th data-sort="number" style="text-align:center">Held by<span class="arrow">⇅</span></th>')
    thead_html.append(
        '<th data-sort="number" class="bench" style="text-align:right" '
        'title="S&amp;P/NZX 50 benchmark weight">NZX50<span class="arrow">⇅</span></th>'
    )
    for fid in fund_ids:
        info = fund_meta.get(fid, {})
        full = info.get("fund_name", fid)
        prov = info.get("provider", "")
        short = make_short_label(full)
        tip = f"{full} — {prov}".replace('"', "&quot;")
        thead_html.append(
            f'<th data-sort="number" style="text-align:right" title="{tip}">'
            f'{html_text(short)}<span class="fid">{fid}</span>'
            f'<span class="arrow">⇅</span></th>'
        )
    thead_html.append('</tr>')
    thead = "".join(thead_html)

    # ── tbody ──────────────────────────────────────────────────────────────
    tbody_rows = []
    for t in sorted_tickers:
        m = meta[t]
        bw = bench.get(t, 0.0)
        fund_weights = [pivot[t].get(fid, 0.0) for fid in fund_ids]
        num_funds_holding = sum(1 for w in fund_weights if w and w > 0.001)

        searchable = f"{t} {m['name']} {m['sector']} {m['country']}"
        tr_attrs = (
            f'data-searchable="{html_attr(searchable)}" '
            f'data-sector="{html_attr(m["sector"])}" '
            f'data-country="{html_attr(m["country"])}" '
            f'data-num-funds-holding="{num_funds_holding}"'
        )
        cells = [
            f'<td class="text ticker" data-v="{t}">{t}</td>',
            f'<td class="text" data-v="{html_attr(m["name"])}">{html_text(m["name"])}</td>',
            f'<td class="text country" data-v="{m["country"]}">{m["country"]}</td>',
            f'<td class="text sector" data-v="{html_attr(m["sector"])}">{html_text(m["sector"])}</td>',
            f'<td class="held-by" data-v="{num_funds_holding}"><strong>{num_funds_holding}</strong>'
            f'<span style="opacity:.5">/{total_funds}</span></td>',
            f'<td class="bench w" data-v="{bw}">{fmt_pct(bw)}</td>',
        ]
        for w in fund_weights:
            classes = ["w", heat_class(w)]
            active = w - bw
            ac = active_class(active)
            if ac: classes.append(ac)
            if not w or abs(w) < 0.005: classes.append("zero")
            classes = [c for c in classes if c]
            cells.append(
                f'<td class="{" ".join(classes)}" data-v="{w}" data-active="{active:.4f}">'
                f'{fmt_pct(w)}</td>'
            )
        tbody_rows.append(f'<tr {tr_attrs}>{"".join(cells)}</tr>')
    tbody = "\n".join(tbody_rows)

    sectors_options = "".join(
        f'<option value="{html_attr(s)}">{html_text(s)}</option>' for s in sectors
    )

    html = (HTML_TEMPLATE
        .replace("__GENERATED__",      datetime.now().strftime("%Y-%m-%d %H:%M"))
        .replace("__FUND_COUNT__",     str(total_funds))
        .replace("__TICKER_COUNT__",   str(total_tickers))
        .replace("__HOLDINGS_COUNT__", str(total_holdings))
        .replace("__THEAD__",          thead)
        .replace("__TBODY__",          tbody)
        .replace("__SECTORS__",        sectors_options))

    with open(DASHBOARD_HTML, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"Wrote {DASHBOARD_HTML}")
    print(f"  {total_tickers} tickers × {total_funds} funds, {total_holdings} holdings")


def make_short_label(fund_name: str) -> str:
    # Strip noisy words and shorten
    words = fund_name.upper().split()
    drop = {"FUND", "FUNDS", "THE", "LIMITED", "PORTFOLIO", "EQUITY", "EQUITIES"}
    keep = [w for w in words if w not in drop]
    if not keep: keep = words[:3]
    short = " ".join(keep)
    if len(short) > 22:
        short = short[:21] + "…"
    return short


def heat_class(w):
    if not w or abs(w) < 0.005: return ""
    a = abs(w)
    if a < 1.0:  return "h1"
    if a < 3.0:  return "h2"
    if a < 6.0:  return "h3"
    if a < 12.0: return "h4"
    return "h5"


def active_class(diff):
    if not diff or abs(diff) < 0.005: return ""
    a = abs(diff)
    side = "pos" if diff > 0 else "neg"
    if a < 1.0:  return f"act-{side}1"
    if a < 3.0:  return f"act-{side}2"
    return f"act-{side}3"


def fmt_pct(w):
    if not w or abs(w) < 0.005:
        return '<span style="color:#d4dae3">—</span>'
    return f"{w:.2f}%"


def html_text(s):
    return (str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"))


def html_attr(s):
    return html_text(s).replace('"', "&quot;")


if __name__ == "__main__":
    render_dashboard()
