#!/usr/bin/env python3
"""
NZ Australasian Funds — Holdings Extractor

For each fund listed in selected_funds.csv:
  1. Download the full portfolio holdings file (xlsx or csv)
  2. Parse the Disclose Register format (auto-detect header row)
  3. Classify each row (equity / cash / fx / settlement / fund / bond)
  4. Normalise security names and resolve to a canonical ticker via
     securities.csv (ISIN-first, name-fallback)
  5. Write three artefacts:
       - data/holdings_clean.csv      long-format clean data, one row per holding
       - data/holdings_matrix.xlsx    wide pivot: equity rows × fund columns
       - data/unmatched.csv           anything we couldn't classify or match
  6. Also writes a benchmark column from the S&P/NZX 50 fund (FND19363) so
     active positions can be computed.

Inputs (kept in repo root, edited by hand):
  - selected_funds.csv   list of funds to process
  - securities.csv       master ticker / canonical name / ISIN / sector list

Outputs (committed back to repo by GitHub Actions):
  - data/holdings_raw/FND*.xlsx    audit trail of downloaded files
  - data/holdings_clean.csv
  - data/holdings_matrix.xlsx
  - data/unmatched.csv
  - data/extract_log.txt
"""

from __future__ import annotations

import csv
import os
import re
import sys
import time
from datetime import datetime
from urllib.parse import urlparse

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.worksheet.page import PageMargins

# ── Config ────────────────────────────────────────────────────────────────────
SCRIPT_DIR    = os.path.dirname(os.path.abspath(__file__))
SELECTED_CSV  = os.path.join(SCRIPT_DIR, "selected_funds.csv")
SECURITIES_CSV= os.path.join(SCRIPT_DIR, "securities.csv")

DATA_DIR      = os.path.join(SCRIPT_DIR, "data")
RAW_DIR       = os.path.join(DATA_DIR, "holdings_raw")
CLEAN_CSV     = os.path.join(DATA_DIR, "holdings_clean.csv")
MATRIX_XLSX   = os.path.join(DATA_DIR, "holdings_matrix.xlsx")
UNMATCHED_CSV = os.path.join(DATA_DIR, "unmatched.csv")
LOG_PATH      = os.path.join(DATA_DIR, "extract_log.txt")

# Benchmark configuration:
#
# To populate the NZX50 benchmark column, add a row to selected_funds.csv with
# tag = "benchmark" pointing at any NZX50-tracking fund. Two good options:
#
#   benchmark,FND1115,SMART NZ TOP 50 ETF,SMARTSHARES LIMITED,...,<xlsx URL>
#   benchmark,FND19363,S&P/NZX 50 FUND,SMARTSHARES LIMITED,...,<xlsx URL>
#
# Get the URL from the fund's Sorted page (the "Complete asset portfolio" link).
# If no benchmark row exists, the matrix will still build but without the
# BENCHMARK column or active-vs-NZX50 calculations.
BENCHMARK_TAG = "benchmark"

# Australasian filter — what counts as "in scope" for the matrix
AUSTRALASIAN_COUNTRIES = {"NZ", "AU"}

DOWNLOAD_DELAY = 0.5  # seconds between downloads — Disclose Register is fine with this

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "*/*",
    "Accept-Language": "en-NZ,en;q=0.9",
}

os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(RAW_DIR,  exist_ok=True)


# ── Logger ───────────────────────────────────────────────────────────────────
_log_lines: list[str] = []

def log(msg: str = "") -> None:
    line = f"[{datetime.now().strftime('%H:%M:%S')}] {msg}" if msg else ""
    print(line)
    _log_lines.append(line)


def flush_log() -> None:
    with open(LOG_PATH, "w", encoding="utf-8") as f:
        f.write("\n".join(_log_lines) + "\n")


# ── Securities master list ───────────────────────────────────────────────────
SUFFIX_TOKENS = {
    "LTD", "LIMITED", "PLC", "INC", "CORPORATION", "CORP",
    "ORDINARY", "SHARES", "STOCK", "COMMON",
    "/THE", "THE",
}
TRAILING_NOISE = {"GROUP", "HOLDINGS", "CO", "COMPANY", "NZ"}


def normalise_name(name: str) -> str:
    """
    Aggressively normalise a security name for fuzzy matching.

    "Fisher & Paykel Healthcare Corp Ltd"  → "FISHER AND PAYKEL HEALTHCARE"
    "FISHER & PAYKEL HEALTHCARE CORP LTD"  → "FISHER AND PAYKEL HEALTHCARE"
    "The A2 Milk Company Limited"          → "A2 MILK"
    "A2 MILK CO LTD/THE"                   → "A2 MILK"
    """
    n = name.upper().strip()
    n = n.replace("&", "AND")
    n = re.sub(r"[/,()]", " ", n)
    n = re.sub(r"[^\w\s]", "", n)
    tokens = [t for t in n.split() if t not in SUFFIX_TOKENS]
    while len(tokens) > 1 and tokens[-1] in TRAILING_NOISE:
        tokens.pop()
    return " ".join(tokens)


class SecurityMaster:
    """Looks up rows by ISIN or by normalised name."""

    def __init__(self, securities_csv: str):
        self._by_isin: dict[str, dict] = {}
        self._by_norm: dict[str, dict] = {}

        with open(securities_csv, encoding="utf-8") as f:
            for row in csv.DictReader(f):
                # skip comment / blank rows
                if not row.get("ticker") or row["ticker"].startswith("#"):
                    continue
                rec = {
                    "ticker":         row["ticker"].strip(),
                    "canonical_name": row["canonical_name"].strip(),
                    "country":        row.get("country", "").strip(),
                    "sector":         row.get("sector", "").strip(),
                    "isin":           row.get("isin", "").strip(),
                }
                if rec["isin"]:
                    self._by_isin[rec["isin"]] = rec

                # Index canonical name + every alias under normalised form
                for alias in [rec["canonical_name"]] + [
                    a.strip() for a in row.get("aliases", "").split("|") if a.strip()
                ]:
                    self._by_norm.setdefault(normalise_name(alias), rec)

        log(f"Loaded {len(self._by_isin)} securities by ISIN, "
            f"{len(self._by_norm)} by name.")

    def lookup(self, name: str, isin: str = "") -> dict | None:
        # ISIN-first (bulletproof)
        if isin and isin in self._by_isin:
            return self._by_isin[isin]
        # Name-fallback
        norm = normalise_name(name)
        if norm in self._by_norm:
            return self._by_norm[norm]
        # Partial match: any normalised key fully contained in our normalised name
        for k, rec in self._by_norm.items():
            if k and k in norm:
                return rec
        return None


# ── Row classifier ───────────────────────────────────────────────────────────
NON_EQUITY_PATTERNS: list[tuple[re.Pattern, str]] = [
    # FX hedges / forwards
    (re.compile(r"\bBUY\s+(NZD|AUD|USD|EUR|GBP)"), "fx"),
    (re.compile(r"\bSELL\s+(NZD|AUD|USD|EUR|GBP)"), "fx"),
    (re.compile(r"^FX\s+"), "fx"),
    (re.compile(r"\bFORWARD\b"), "fx"),
    # Cash / current accounts / deposits
    (re.compile(r"\bCURRENT\s+ACCOUNT"), "cash"),
    (re.compile(r"\bDEPOSIT\s+ACCOUNT"), "cash"),
    (re.compile(r"\bCASH\s+ACCOUNT"), "cash"),
    (re.compile(r"\bWHOLESALE\s+CASH"), "cash"),
    (re.compile(r"\bMONEY\s+MARKET"), "cash"),
    (re.compile(r"\bA/?C\b(?!\w)"), "cash"),     # "A/C" or "AC" but not "ACCO"...
    (re.compile(r"\bMARGIN\s+(AC|ACCOUNT)"), "cash"),
    (re.compile(r"\bCALL\s+ACCOUNT"), "cash"),
    (re.compile(r"^CASH\b"), "cash"),
    (re.compile(r"\bCASH\s*$"), "cash"),
    # Settlement / clearing
    (re.compile(r"\bSETTLEMENT"), "settlement"),
    (re.compile(r"\bNZCLEAR"), "settlement"),
    (re.compile(r"\bNET\s+CURRENT\s+ASSETS"), "settlement"),
    # Bonds / fixed interest
    (re.compile(r"\b\d+(\.\d+)?\s*%\s+\d{4}"), "bond"),  # "5.5% 2030"
    (re.compile(r"\bFIXED\s+INTEREST"), "bond"),
    (re.compile(r"\bBOND\b"), "bond"),
    (re.compile(r"\bDEBENTURE"), "bond"),
    # Look-through to other funds — these are flagged so user can decide whether
    # to drill into them. We never silently include them as equities.
    (re.compile(r"\bUNIT\s+TRUST"), "fund"),
    (re.compile(r"\bETF\s*$"), "fund"),
    (re.compile(r"\bETF\s+UNITS"), "fund"),
    (re.compile(r"\bWHOLESALE\s+FUND"), "fund"),
    (re.compile(r"^iShares\b", re.I), "fund"),
    (re.compile(r"^Vanguard\b", re.I), "fund"),
    (re.compile(r"^SmartShares\b", re.I), "fund"),
    (re.compile(r"\bFund\s+units\b", re.I), "fund"),
]

ISIN_RE = re.compile(r"^([A-Z]{2})[A-Z0-9]{9}\d$")


def classify_row(name: str, isin: str = "") -> str:
    """
    Returns one of: equity, cash, fx, settlement, bond, fund, other.
    'equity' rows continue to the matrix; everything else is logged & excluded.
    """
    name_upper = name.upper().strip()
    if not name_upper:
        return "other"

    # If any non-equity pattern fires, that wins (even if we have an ISIN —
    # some funds list their cash accounts with funny near-ISIN-looking codes).
    for pattern, tag in NON_EQUITY_PATTERNS:
        if pattern.search(name_upper):
            return tag

    # Has a valid ISIN that's NZ or AU? Definitely an equity.
    isin_match = ISIN_RE.match(isin.strip()) if isin else None
    if isin_match and isin_match.group(1) in AUSTRALASIAN_COUNTRIES:
        return "equity"

    # Has a valid non-Australasian ISIN? Still an equity, but we'll filter it
    # out at the matrix stage by country.
    if isin_match:
        return "equity"

    # No ISIN and no non-equity match — assume equity. Will be flagged for
    # review if it can't be resolved against the master list.
    return "equity"


# ── File parser ──────────────────────────────────────────────────────────────
def _read_csv_robust(path: str) -> list[list[str]]:
    """
    Read a CSV file with best-effort encoding detection. Disclose Register CSVs
    are sometimes UTF-8, sometimes UTF-8 with BOM, and occasionally Windows-1252
    (legacy Excel exports). Try them in order; fall back to Latin-1 (which can
    decode any byte) so we never fail with UnicodeDecodeError.
    """
    encodings = ("utf-8-sig", "utf-8", "cp1252", "latin-1")
    for enc in encodings:
        try:
            with open(path, encoding=enc) as f:
                return [[c.strip() for c in row] for row in csv.reader(f)]
        except UnicodeDecodeError:
            continue
    # Latin-1 above can never raise UnicodeDecodeError, so we should never
    # reach here. Defensive fallback regardless.
    with open(path, encoding="latin-1", errors="replace") as f:
        return [[c.strip() for c in row] for row in csv.reader(f)]


def _detect_format(path: str) -> str:
    """
    Return 'xlsx' or 'csv' based on file content (not extension).
    Some Disclose Register URLs ending in .csv actually serve xlsx binary —
    we need to handle that gracefully rather than trusting the URL.
    """
    with open(path, "rb") as f:
        head = f.read(4)
    # xlsx (and other Office 2007+ formats) are ZIP archives — magic is "PK\x03\x04"
    if head[:2] == b"PK":
        return "xlsx"
    # Old .xls (Office 97-2003) magic — we don't support these but flag clearly
    if head[:4] == b"\xd0\xcf\x11\xe0":
        return "xls"
    return "csv"


def read_holdings_file(path: str) -> list[dict]:
    """
    Return a list of {asset_name, weight_pct, isin, source_row} dicts.
    Handles both .xlsx and .csv Disclose Register exports. Auto-detects the
    header row by scanning for "asset name" in column A.

    Detects format by content (magic bytes), not file extension — some servers
    serve xlsx content under a .csv URL.
    """
    rows: list[list[str]]

    fmt = _detect_format(path)
    if fmt == "xls":
        raise ValueError(
            f"{path}: legacy .xls (Office 97-2003) format not supported. "
            f"Open in Excel and re-save as .xlsx, or download the .xlsx version."
        )

    if fmt == "xlsx":
        # Read bytes and pass to openpyxl as a BytesIO so it doesn't reject
        # files with the wrong extension (e.g. xlsx content served under .csv URL).
        from io import BytesIO
        with open(path, "rb") as f:
            data = f.read()
        wb = load_workbook(BytesIO(data), data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = []
        max_col = max(4, ws.max_column)
        for r in range(1, ws.max_row + 1):
            row = [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
            rows.append(["" if v is None else str(v).strip() for v in row])
    else:  # csv
        rows = _read_csv_robust(path)

    # Find header row (col A contains "asset name")
    header_idx = None
    for i, row in enumerate(rows):
        if row and "asset name" in row[0].lower():
            header_idx = i
            break
    if header_idx is None:
        raise ValueError(f"No 'Asset name' header row found in {path}")

    out: list[dict] = []
    for i, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
        # pad row to at least 3 cols
        row = row + [""] * (3 - len(row))
        name = row[0].strip()
        if not name:
            continue
        wt_raw = row[1].rstrip("%").strip() if len(row) > 1 else ""
        if not wt_raw:
            continue
        try:
            weight = float(wt_raw)
        except ValueError:
            continue
        isin = row[2].strip() if len(row) > 2 else ""
        out.append({
            "asset_name":  name,
            "weight_pct":  weight,
            "isin":        isin,
            "source_row":  i,
        })
    return out


# ── Downloader ───────────────────────────────────────────────────────────────
def filename_for_url(fund_id: str, url: str) -> str:
    parsed = urlparse(url)
    suffix = ".xlsx" if parsed.path.lower().endswith(".xlsx") else (
             ".csv"  if parsed.path.lower().endswith(".csv") else
             os.path.splitext(parsed.path)[1] or ".bin")
    return os.path.join(RAW_DIR, f"{fund_id}{suffix}")


def find_cached(fund_id: str) -> str | None:
    """Return path to a cached holdings file for fund_id, regardless of extension."""
    for ext in (".xlsx", ".csv"):
        p = os.path.join(RAW_DIR, f"{fund_id}{ext}")
        if os.path.exists(p) and os.path.getsize(p) > 1000:
            return p
    return None


def download(fund_id: str, url: str, force: bool = False) -> str | None:
    """Download to RAW_DIR. Returns local path or None on failure.
    Detects xlsx-served-as-csv (or vice versa) and renames the file to match."""
    if not url:
        log(f"  {fund_id}: no URL given — skipping")
        return None

    # Cache hit on any extension (xlsx or csv)
    cached = find_cached(fund_id)
    if cached and not force:
        log(f"  {fund_id}: cached at {os.path.basename(cached)}")
        return cached

    local = filename_for_url(fund_id, url)
    log(f"  {fund_id}: downloading {url}")
    try:
        r = requests.get(url, headers=HEADERS, timeout=60)
        r.raise_for_status()
    except Exception as e:
        log(f"  {fund_id}: FAILED — {e}")
        return None

    with open(local, "wb") as f:
        f.write(r.content)

    # Sniff content: server may serve xlsx under a .csv URL (or vice versa).
    # Rename file to match the actual content so caching/parsing don't get
    # confused on subsequent runs.
    actual_fmt = _detect_format(local)
    expected_ext = ".xlsx" if actual_fmt == "xlsx" else ".csv"
    if not local.lower().endswith(expected_ext):
        new_local = os.path.join(RAW_DIR, f"{fund_id}{expected_ext}")
        os.replace(local, new_local)
        log(f"  {fund_id}: server returned {actual_fmt} content — "
            f"renamed to {os.path.basename(new_local)}")
        local = new_local

    log(f"  {fund_id}: saved {len(r.content):,} bytes → {os.path.basename(local)}")
    time.sleep(DOWNLOAD_DELAY)
    return local


# ── Reader for selected_funds.csv ────────────────────────────────────────────
def read_selected_funds() -> list[dict]:
    funds = []
    with open(SELECTED_CSV, encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            if not row.get("fund_id"):
                continue
            funds.append({
                "tag":        row.get("tag", "").strip(),
                "fund_id":    row["fund_id"].strip(),
                "fund_name":  row.get("fund_name", "").strip(),
                "provider":   row.get("provider", "").strip(),
                "scheme":     row.get("scheme", "").strip(),
                "fund_type":  row.get("fund_type", "").strip(),
                "fees_pct":   row.get("fees_pct", "").strip(),
                "url":        row.get("portfolio_xlsx_url", "").strip(),
            })
    # Split out any benchmark rows so they don't appear as fund columns
    benchmark_rows = [f for f in funds if f["tag"].lower() == BENCHMARK_TAG]
    regular_funds  = [f for f in funds if f["tag"].lower() != BENCHMARK_TAG]
    log(f"Loaded {len(regular_funds)} selected funds + "
        f"{len(benchmark_rows)} benchmark fund(s) from {os.path.basename(SELECTED_CSV)}")
    return regular_funds, benchmark_rows


# (read_selected_funds returns a tuple now — earlier callers must unpack)


# ── Per-fund processing ──────────────────────────────────────────────────────
def process_fund(
    fund: dict,
    sec_master: SecurityMaster,
    rows_out: list[dict],
    unmatched_out: list[dict],
) -> dict:
    """
    Process a single fund. Appends rows to the shared lists. Returns a summary
    dict with counts/totals for the per-fund report.
    """
    fund_id   = fund["fund_id"]
    fund_name = fund["fund_name"]
    summary = {
        "fund_id":              fund_id,
        "fund_name":            fund_name,
        "provider":             fund["provider"],
        "fees_pct":             fund["fees_pct"],
        "downloaded":           False,
        "raw_rows":             0,
        "equity_rows":          0,
        "australasian_rows":    0,
        "matched_rows":         0,
        "unmatched_rows":       0,
        "australasian_weight":  0.0,
        "non_equity_weight":    0.0,
        "non_australasian_weight": 0.0,
    }

    local = download(fund_id, fund["url"])
    if not local:
        return summary
    summary["downloaded"] = True

    try:
        raw_rows = read_holdings_file(local)
    except Exception as e:
        log(f"  {fund_id}: parse FAILED — {e}")
        return summary

    summary["raw_rows"] = len(raw_rows)
    log(f"  {fund_id}: parsed {len(raw_rows)} rows; classifying & matching…")

    for r in raw_rows:
        cls = classify_row(r["asset_name"], r["isin"])

        if cls != "equity":
            summary["non_equity_weight"] += r["weight_pct"]
            unmatched_out.append({
                "fund_id":     fund_id,
                "fund_name":   fund_name,
                "asset_name":  r["asset_name"],
                "weight_pct":  r["weight_pct"],
                "isin":        r["isin"],
                "reason":      f"classified as {cls}",
            })
            continue

        summary["equity_rows"] += 1
        rec = sec_master.lookup(r["asset_name"], r["isin"])

        if rec is None:
            # Equity-shaped but not in our master list. Maybe an
            # international stock, maybe a name we haven't aliased yet.
            summary["unmatched_rows"] += 1
            unmatched_out.append({
                "fund_id":     fund_id,
                "fund_name":   fund_name,
                "asset_name":  r["asset_name"],
                "weight_pct":  r["weight_pct"],
                "isin":        r["isin"],
                "reason":      "no master match",
            })
            continue

        country = rec["country"]
        if country not in AUSTRALASIAN_COUNTRIES:
            summary["non_australasian_weight"] += r["weight_pct"]
            continue

        summary["matched_rows"]        += 1
        summary["australasian_rows"]   += 1
        summary["australasian_weight"] += r["weight_pct"]
        rows_out.append({
            "fund_id":         fund_id,
            "fund_name":       fund_name,
            "asset_name":      r["asset_name"],
            "ticker":          rec["ticker"],
            "canonical_name":  rec["canonical_name"],
            "country":         rec["country"],
            "sector":          rec["sector"],
            "isin_raw":        r["isin"],
            "isin_master":     rec["isin"],
            "weight_pct":      round(r["weight_pct"], 4),
        })

    log(f"  {fund_id}: {summary['australasian_rows']} Australasian equities "
        f"({summary['australasian_weight']:.1f}% of fund), "
        f"{summary['unmatched_rows']} unmatched, "
        f"{summary['non_equity_weight']:.1f}% non-equity")
    return summary


# ── Output: long-format CSV ──────────────────────────────────────────────────
def write_clean_csv(rows: list[dict]) -> None:
    cols = [
        "fund_id", "fund_name", "ticker", "canonical_name",
        "country", "sector", "weight_pct",
        "asset_name", "isin_raw", "isin_master",
    ]
    with open(CLEAN_CSV, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=cols, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow(r)
    log(f"Wrote {CLEAN_CSV} ({len(rows)} rows)")


def write_unmatched_csv(rows: list[dict]) -> None:
    cols = ["fund_id", "fund_name", "asset_name", "weight_pct", "isin", "reason"]
    with open(UNMATCHED_CSV, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=cols, extrasaction="ignore")
        w.writeheader()
        # Sort: most material rows first
        for r in sorted(rows, key=lambda x: -abs(x.get("weight_pct", 0) or 0)):
            w.writerow(r)
    log(f"Wrote {UNMATCHED_CSV} ({len(rows)} rows)")


# ── Output: matrix XLSX ──────────────────────────────────────────────────────
def write_matrix_xlsx(
    rows: list[dict],
    funds: list[dict],
    summaries: list[dict],
    benchmark_weights: dict[str, float],
) -> None:
    """
    Build the wide-format peer-fund matrix:

      Sheet "Matrix":
        rows    = unique tickers (sorted by total weight across funds, desc)
        cols    = each selected fund + benchmark + active vs benchmark per fund
        cells   = % weight

      Sheet "Long":
        flat one-row-per-holding view

      Sheet "Summary":
        per-fund stats — coverage, weight totals, unmatched count

      Sheet "By Sector":
        same matrix grouped/sorted by sector
    """
    # Build pivot
    fund_ids   = [f["fund_id"] for f in funds]
    fund_names = {f["fund_id"]: f["fund_name"] for f in funds}

    # ticker -> fund_id -> weight
    pivot: dict[str, dict[str, float]] = {}
    ticker_meta: dict[str, dict] = {}  # ticker -> {canonical_name, country, sector}

    for r in rows:
        t = r["ticker"]
        pivot.setdefault(t, {})[r["fund_id"]] = (
            pivot.setdefault(t, {}).get(r["fund_id"], 0.0) + r["weight_pct"]
        )
        ticker_meta[t] = {
            "canonical_name": r["canonical_name"],
            "country":        r["country"],
            "sector":         r["sector"],
        }

    # Add benchmark tickers we don't already have
    for t, w in benchmark_weights.items():
        if t not in ticker_meta:
            # We have it in benchmark but no fund holds it — skip; matrix only
            # cares about tickers that appear in at least one selected fund.
            pass

    # Sort tickers by SUM weight across funds (most-held first)
    def total_weight(t: str) -> float:
        return sum(pivot.get(t, {}).values())
    sorted_tickers = sorted(pivot.keys(), key=lambda t: -total_weight(t))

    wb = Workbook()

    # ── Sheet 1: Matrix ───────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Matrix"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2C3E50")
    sub_fill    = PatternFill("solid", fgColor="34495E")

    base_cols = ["Ticker", "Name", "Country", "Sector"]
    fund_cols = fund_ids
    bench_col = "BENCHMARK"
    active_cols = [f"{fid}_active" for fid in fund_ids]
    all_cols = base_cols + fund_cols + [bench_col] + active_cols

    # Two-row header: top row = group labels, second row = column codes
    # For now, single-row header with merged sections
    for col_idx, h in enumerate(base_cols, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.font = header_font; c.fill = header_fill
    for j, fid in enumerate(fund_cols):
        c = ws.cell(row=1, column=len(base_cols) + 1 + j, value=fid)
        c.font = header_font; c.fill = header_fill
    bench_col_idx = len(base_cols) + len(fund_cols) + 1
    c = ws.cell(row=1, column=bench_col_idx, value=bench_col)
    c.font = header_font; c.fill = PatternFill("solid", fgColor="16A085")
    for j, ac in enumerate(active_cols):
        col_idx = bench_col_idx + 1 + j
        c = ws.cell(row=1, column=col_idx, value=ac)
        c.font = header_font; c.fill = sub_fill

    # Data rows
    for row_idx, ticker in enumerate(sorted_tickers, start=2):
        meta = ticker_meta[ticker]
        ws.cell(row=row_idx, column=1, value=ticker).font = Font(bold=True)
        ws.cell(row=row_idx, column=2, value=meta["canonical_name"])
        ws.cell(row=row_idx, column=3, value=meta["country"])
        ws.cell(row=row_idx, column=4, value=meta["sector"])

        bench_w = benchmark_weights.get(ticker, 0.0)
        ws.cell(row=row_idx, column=bench_col_idx, value=bench_w if bench_w else None)\
            .number_format = '0.00"%"'

        for j, fid in enumerate(fund_cols):
            w = pivot.get(ticker, {}).get(fid, 0.0)
            cell = ws.cell(row=row_idx, column=len(base_cols) + 1 + j,
                           value=w if w else None)
            cell.number_format = '0.00"%"'

            active = (w if w else 0.0) - (bench_w if bench_w else 0.0)
            ac_cell = ws.cell(
                row=row_idx,
                column=bench_col_idx + 1 + j,
                value=active if w or bench_w else None,
            )
            ac_cell.number_format = '+0.00"%";-0.00"%"'

    # Column widths
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 16
    for col_idx in range(len(base_cols) + 1, len(all_cols) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 12
    ws.freeze_panes = "E2"

    # ── Conditional formatting (heat map on the data area) ──────────────────
    last_row = len(sorted_tickers) + 1
    if last_row >= 2:
        # Weight cells: light → dark blue scale (0% → 20%)
        # Apply across the fund columns only (NOT the active columns)
        weight_first = get_column_letter(len(base_cols) + 1)            # E
        weight_last  = get_column_letter(bench_col_idx)                 # benchmark col
        weight_range = f"{weight_first}2:{weight_last}{last_row}"
        ws.conditional_formatting.add(
            weight_range,
            ColorScaleRule(
                start_type="num", start_value=0,    start_color="FFFFFF",
                mid_type="num",   mid_value=5,      mid_color="A8CAEE",
                end_type="num",   end_value=20,     end_color="2C3E50",
            ),
        )

        # Active cells: red (negative) ↔ green (positive) divergent scale
        if active_cols:
            active_first = get_column_letter(bench_col_idx + 1)
            active_last  = get_column_letter(bench_col_idx + len(active_cols))
            active_range = f"{active_first}2:{active_last}{last_row}"
            ws.conditional_formatting.add(
                active_range,
                ColorScaleRule(
                    start_type="num", start_value=-8,  start_color="C0392B",
                    mid_type="num",   mid_value=0,     mid_color="FFFFFF",
                    end_type="num",   end_value=8,     end_color="1E7C4A",
                ),
            )

    # ── Print setup ─────────────────────────────────────────────────────────
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize   = ws.PAPERSIZE_A3
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0     # 0 = auto / unlimited rows
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.gridLines = False
    ws.print_title_rows       = "1:1"   # repeat header on every printed page
    ws.print_title_cols       = "A:B"   # repeat ticker + name on each page
    ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.5, bottom=0.5,
                                  header=0.2, footer=0.2)

    log(f"  Matrix sheet: {len(sorted_tickers)} rows × {len(fund_cols)} funds")

    # ── Sheet 2: By Sector ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("By Sector")
    # Same columns as Matrix but grouped by sector
    for col_idx, h in enumerate(base_cols + fund_cols + [bench_col] + active_cols, start=1):
        c = ws2.cell(row=1, column=col_idx, value=h)
        c.font = header_font
        c.fill = header_fill if col_idx <= len(base_cols) else (
                 PatternFill("solid", fgColor="16A085") if col_idx == bench_col_idx else sub_fill)

    # Group tickers by sector
    by_sector: dict[str, list[str]] = {}
    for t in sorted_tickers:
        s = ticker_meta[t]["sector"] or "Other"
        by_sector.setdefault(s, []).append(t)

    out_row = 2
    for sector in sorted(by_sector.keys()):
        # Section header
        ws2.cell(row=out_row, column=1, value=sector).font = Font(bold=True, italic=True)
        out_row += 1
        for ticker in by_sector[sector]:
            meta = ticker_meta[ticker]
            ws2.cell(row=out_row, column=1, value=ticker).font = Font(bold=True)
            ws2.cell(row=out_row, column=2, value=meta["canonical_name"])
            ws2.cell(row=out_row, column=3, value=meta["country"])
            ws2.cell(row=out_row, column=4, value=meta["sector"])
            bench_w = benchmark_weights.get(ticker, 0.0)
            ws2.cell(row=out_row, column=bench_col_idx, value=bench_w if bench_w else None)\
                .number_format = '0.00"%"'
            for j, fid in enumerate(fund_cols):
                w = pivot.get(ticker, {}).get(fid, 0.0)
                ws2.cell(row=out_row, column=len(base_cols) + 1 + j,
                         value=w if w else None).number_format = '0.00"%"'
                active = (w if w else 0) - (bench_w if bench_w else 0)
                ws2.cell(row=out_row, column=bench_col_idx + 1 + j,
                         value=active if w or bench_w else None
                ).number_format = '+0.00"%";-0.00"%"'
            out_row += 1

    ws2.column_dimensions["A"].width = 8
    ws2.column_dimensions["B"].width = 38
    ws2.column_dimensions["C"].width = 8
    ws2.column_dimensions["D"].width = 16
    for col_idx in range(len(base_cols) + 1, len(all_cols) + 1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = 12
    ws2.freeze_panes = "E2"

    # Same conditional-formatting heat map on the By Sector sheet
    last_row_2 = out_row - 1
    if last_row_2 >= 2:
        weight_first = get_column_letter(len(base_cols) + 1)
        weight_last  = get_column_letter(bench_col_idx)
        ws2.conditional_formatting.add(
            f"{weight_first}2:{weight_last}{last_row_2}",
            ColorScaleRule(
                start_type="num", start_value=0,  start_color="FFFFFF",
                mid_type="num",   mid_value=5,    mid_color="A8CAEE",
                end_type="num",   end_value=20,   end_color="2C3E50",
            ),
        )
        if active_cols:
            af = get_column_letter(bench_col_idx + 1)
            al = get_column_letter(bench_col_idx + len(active_cols))
            ws2.conditional_formatting.add(
                f"{af}2:{al}{last_row_2}",
                ColorScaleRule(
                    start_type="num", start_value=-8, start_color="C0392B",
                    mid_type="num",   mid_value=0,    mid_color="FFFFFF",
                    end_type="num",   end_value=8,    end_color="1E7C4A",
                ),
            )

    ws2.page_setup.orientation = ws2.ORIENTATION_LANDSCAPE
    ws2.page_setup.paperSize   = ws2.PAPERSIZE_A3
    ws2.page_setup.fitToWidth  = 1
    ws2.page_setup.fitToHeight = 0
    ws2.sheet_properties.pageSetUpPr.fitToPage = True
    ws2.print_options.gridLines = False
    ws2.print_title_rows       = "1:1"

    # ── Sheet 3: Long ──────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Long")
    long_cols = [
        "fund_id", "fund_name", "ticker", "canonical_name",
        "country", "sector", "weight_pct",
        "asset_name", "isin_raw", "isin_master",
    ]
    for col_idx, h in enumerate(long_cols, start=1):
        c = ws3.cell(row=1, column=col_idx, value=h)
        c.font = header_font; c.fill = header_fill
    for r_idx, r in enumerate(rows, start=2):
        for col_idx, k in enumerate(long_cols, start=1):
            v = r.get(k)
            cell = ws3.cell(row=r_idx, column=col_idx, value=v if v != "" else None)
            if k == "weight_pct" and isinstance(v, (int, float)):
                cell.number_format = '0.00"%"'

    last_col = get_column_letter(len(long_cols))
    last_row = len(rows) + 1
    if last_row >= 2:
        tbl = Table(displayName="LongHoldings", ref=f"A1:{last_col}{last_row}")
        tbl.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showRowStripes=True,
        )
        ws3.add_table(tbl)
    ws3.freeze_panes = "A2"
    for col_idx, w in enumerate([10, 35, 8, 32, 8, 18, 12, 38, 16, 16], start=1):
        ws3.column_dimensions[get_column_letter(col_idx)].width = w

    # ── Sheet 4: Summary ────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Summary")
    summary_cols = [
        ("fund_id",                "Fund ID",                14),
        ("fund_name",              "Fund Name",              42),
        ("provider",               "Provider",               34),
        ("fees_pct",               "Fees %",                 9),
        ("raw_rows",               "Raw rows",               10),
        ("equity_rows",            "Equity rows",            12),
        ("australasian_rows",      "Aust'asian eq.",         15),
        ("matched_rows",           "Matched",                10),
        ("unmatched_rows",         "Unmatched",              11),
        ("australasian_weight",    "Aust'asian %",           14),
        ("non_australasian_weight","Non-Aust'asian %",       18),
        ("non_equity_weight",      "Non-equity %",           14),
    ]
    for col_idx, (_, h, w) in enumerate(summary_cols, start=1):
        c = ws4.cell(row=1, column=col_idx, value=h)
        c.font = header_font; c.fill = header_fill
        ws4.column_dimensions[get_column_letter(col_idx)].width = w

    for r_idx, s in enumerate(summaries, start=2):
        for col_idx, (k, _, _) in enumerate(summary_cols, start=1):
            v = s.get(k, "")
            cell = ws4.cell(row=r_idx, column=col_idx, value=v if v != "" else None)
            if k.endswith("_weight") and isinstance(v, (int, float)):
                cell.number_format = '0.00"%"'
    ws4.freeze_panes = "A2"

    # ── README sheet ───────────────────────────────────────────────────────
    ws5 = wb.create_sheet("README")
    readme = [
        ("Generated",        datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Funds",            len(funds)),
        ("Tickers in matrix",len(sorted_tickers)),
        ("Long-format rows", len(rows)),
        ("",                 ""),
        ("Sheet: Matrix",    "Tickers as rows, funds as columns, % weights as cells. "
                             "Active columns = fund weight − benchmark weight."),
        ("Sheet: By Sector", "Same data, grouped by sector for thematic scanning."),
        ("Sheet: Long",      "One row per fund×holding. Use this for filtering / pivoting."),
        ("Sheet: Summary",   "Per-fund stats: coverage, weight totals, unmatched count."),
        ("",                 ""),
        ("Workflow",         "Edit selected_funds.csv to add/remove funds, "
                             "edit securities.csv to add aliases for unmatched names, "
                             "then re-run the GitHub Actions workflow."),
    ]
    for r_idx, (k, v) in enumerate(readme, start=1):
        ws5.cell(row=r_idx, column=1, value=k).font = Font(bold=True)
        ws5.cell(row=r_idx, column=2, value=v).alignment = Alignment(wrap_text=True, vertical="top")
    ws5.column_dimensions["A"].width = 22
    ws5.column_dimensions["B"].width = 80

    wb.save(MATRIX_XLSX)
    log(f"Wrote {MATRIX_XLSX}")


# ── Benchmark loader ─────────────────────────────────────────────────────────
def load_benchmark(
    benchmark_rows: list[dict], sec_master: SecurityMaster
) -> dict[str, float]:
    """
    Take any rows tagged "benchmark" in selected_funds.csv, download and parse
    them, and return {ticker: weight_pct}. If multiple benchmark rows exist,
    weights are averaged across them.

    Failures here are NOT fatal — the rest of the matrix still works without
    a benchmark column.
    """
    if not benchmark_rows:
        log("No benchmark row found in selected_funds.csv "
            "(tag a row with 'benchmark' to enable NZX50 active comparison)")
        return {}

    log(f"Fetching benchmark from {len(benchmark_rows)} row(s)…")
    all_weights: list[dict[str, float]] = []
    for bench in benchmark_rows:
        fund_id = bench["fund_id"]
        log(f"  benchmark {fund_id}: {bench['fund_name']}")
        try:
            local = download(fund_id, bench["url"])
        except Exception as e:
            log(f"    download crashed — {e}")
            continue
        if not local:
            log(f"    download failed — skipping")
            continue
        try:
            raw = read_holdings_file(local)
        except Exception as e:
            log(f"    parse failed — {e}")
            continue

        weights: dict[str, float] = {}
        for r in raw:
            if classify_row(r["asset_name"], r["isin"]) != "equity":
                continue
            rec = sec_master.lookup(r["asset_name"], r["isin"])
            if not rec:
                log(f"    benchmark unmatched: {r['asset_name']} ({r['isin']})")
                continue
            weights[rec["ticker"]] = weights.get(rec["ticker"], 0.0) + r["weight_pct"]
        log(f"    parsed {len(weights)} tickers")
        all_weights.append(weights)

    if not all_weights:
        log("  All benchmark sources failed — continuing without benchmark column.")
        return {}

    # Average across benchmark sources (usually just one)
    if len(all_weights) == 1:
        return all_weights[0]
    combined: dict[str, float] = {}
    for ws in all_weights:
        for t, w in ws.items():
            combined[t] = combined.get(t, 0.0) + w
    n = len(all_weights)
    return {t: w / n for t, w in combined.items()}


# ── Entry point ──────────────────────────────────────────────────────────────
def main() -> int:
    started = datetime.now()
    log(f"NZ Funds Holdings Extractor — {started.isoformat(timespec='seconds')}")
    log("")

    sec_master = SecurityMaster(SECURITIES_CSV)
    funds, benchmark_rows = read_selected_funds()

    benchmark_weights = load_benchmark(benchmark_rows, sec_master)
    log("")

    rows: list[dict]      = []
    unmatched: list[dict] = []
    summaries: list[dict] = []

    for fund in funds:
        log(f"Fund {fund['fund_id']}: {fund['fund_name']}")
        s = process_fund(fund, sec_master, rows, unmatched)
        summaries.append(s)

    log("")
    log("Writing outputs…")
    write_clean_csv(rows)
    write_unmatched_csv(unmatched)
    write_matrix_xlsx(rows, funds, summaries, benchmark_weights)
    flush_log()

    elapsed = datetime.now() - started
    log(f"Done in {elapsed}.")
    flush_log()
    return 0


if __name__ == "__main__":
    sys.exit(main())
